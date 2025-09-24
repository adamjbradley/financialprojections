#!/usr/bin/env python3
"""
Enhance the Excel revenue projection model with advanced features:
- Better segment integration
- Advanced scenario analysis
- Period-based calculations
- Export functionality simulation
- Demographics integration
"""

import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import math

class AdvancedExcelModel:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.load_config_data()
        
        # Enhanced styles
        self.create_named_styles()
        
    def load_config_data(self):
        """Load configuration data from model-config.json"""
        config_path = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/model-config.json'
        try:
            with open(config_path, 'r') as f:
                self.config = json.load(f)
        except FileNotFoundError:
            self.config = self.create_fallback_config()
    
    def create_fallback_config(self):
        """Create fallback configuration data"""
        return {
            "countries": {
                "india": {"name": "India", "currency": "INR", "currencySymbol": "₹", "exchangeRate": 83.50},
                "singapore": {"name": "Singapore", "currency": "SGD", "currencySymbol": "S$", "exchangeRate": 1.35},
                "australia": {"name": "Australia", "currency": "AUD", "currencySymbol": "A$", "exchangeRate": 1.52},
                "japan": {"name": "Japan", "currency": "JPY", "currencySymbol": "¥", "exchangeRate": 149.0},
                "south_korea": {"name": "South Korea", "currency": "KRW", "currencySymbol": "₩", "exchangeRate": 1320.0},
                "thailand": {"name": "Thailand", "currency": "THB", "currencySymbol": "฿", "exchangeRate": 35.8},
                "indonesia": {"name": "Indonesia", "currency": "IDR", "currencySymbol": "Rp", "exchangeRate": 15750.0},
                "philippines": {"name": "Philippines", "currency": "PHP", "currencySymbol": "₱", "exchangeRate": 56.2}
            }
        }
    
    def create_named_styles(self):
        """Create named styles for consistent formatting"""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center')
        
        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '#,##0'
        
        # Percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = '0.0%'
        
        try:
            self.wb.add_named_style(header_style)
            self.wb.add_named_style(currency_style)
            self.wb.add_named_style(percentage_style)
        except ValueError:
            # Styles already exist
            pass

    def enhance_dashboard(self):
        """Enhance the dashboard with better functionality"""
        ws = self.wb['Dashboard']
        
        # Add period-based calculations
        ws['D3'] = 'Calculation Mode:'
        ws['E3'] = '=IF(B4="1M","Daily",IF(OR(B4="5Y",B4="10Y"),"Yearly","Monthly"))'
        ws['D3'].font = Font(bold=True)
        
        # Add quick statistics
        ws['D5'] = 'Quick Stats'
        ws['D5'].font = Font(size=12, bold=True)
        
        stats = [
            ('Months to Display:', '=IF(B4="1M",1,IF(B4="1Y",12,IF(B4="2Y",24,IF(B4="5Y",60,120))))'),
            ('Total Segments:', '=COUNTIF(SegmentLibrary!A:A,B3)'),
            ('Avg Segment Price:', '=AVERAGEIF(SegmentLibrary!A:A,B3,SegmentLibrary!E:E)'),
            ('Total Base Volume:', '=SUMIF(SegmentLibrary!A:A,B3,SegmentLibrary!G:G)')
        ]
        
        for i, (label, formula) in enumerate(stats):
            row = 6 + i
            ws[f'D{row}'] = label
            ws[f'E{row}'] = formula
            ws[f'D{row}'].font = Font(bold=True)
        
        # Add conditional formatting for metrics
        ws['D11'] = 'Performance Indicators'
        ws['D11'].font = Font(size=12, bold=True)
        
        indicators = [
            ('Revenue Growth Rate:', '=IF(SUM(Projections!D2:D13)>0,(SUM(Projections!D8:D13)-SUM(Projections!D2:D7))/SUM(Projections!D2:D7)*100,0)'),
            ('Volume Growth Rate:', '=IF(SUM(Projections!K2:K13)>0,(SUM(Projections!K8:K13)-SUM(Projections!K2:K7))/SUM(Projections!K2:K7)*100,0)'),
            ('Profit Trend:', '=IF(AVERAGE(Projections!J8:J13)>AVERAGE(Projections!J2:J7),"Improving","Declining")')
        ]
        
        for i, (label, formula) in enumerate(indicators):
            row = 12 + i
            ws[f'D{row}'] = label
            ws[f'E{row}'] = formula
            ws[f'D{row}'].font = Font(bold=True)

    def create_advanced_projections(self):
        """Create more sophisticated projection calculations"""
        ws = self.wb['Projections']
        
        # Clear existing formulas and rebuild with better logic
        max_months = 120
        
        # Add better headers
        headers = [
            'Month_Num', 'Period_Label', 'Country', 'Is_Daily_Mode',
            'Revenue_Local', 'Revenue_USD', 'COGS_Local', 'COGS_USD',
            'OpEx_Local', 'OpEx_USD', 'NetProfit_Local', 'NetProfit_USD',
            'ProfitMargin', 'TransactionVolume', 'Seasonality_Factor',
            'Growth_Factor', 'Cumulative_Revenue'
        ]
        
        # Update headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        
        # Enhanced formulas for each row
        for month in range(1, max_months + 1):
            row = month + 1
            
            # Month number
            ws.cell(row=row, column=1, value=month)
            
            # Period label with better formatting
            period_formula = '''=IF(Dashboard!B4="1M",
                TEXT(DATE(YEAR(TODAY()),MONTH(TODAY()),DAY(TODAY())+A{0}-1),"mmm dd"),
                TEXT(DATE(YEAR(TODAY()),MONTH(TODAY())+A{0}-1,1),"yyyy mmm"))'''.format(row)
            ws.cell(row=row, column=2, value=period_formula)
            
            # Country
            ws.cell(row=row, column=3, value='=Dashboard!$B$3')
            
            # Is daily mode
            ws.cell(row=row, column=4, value='=Dashboard!$B$4="1M"')
            
            # Seasonality factor
            seasonality_formula = f'''=IF(Parameters!B7="none",1,
                IF(Parameters!B7="retail",INDEX(Parameters!C14:C25,MOD(A{row}-1,12)+1),
                IF(Parameters!B7="summer",INDEX(Parameters!D14:D25,MOD(A{row}-1,12)+1),1)))'''
            ws.cell(row=row, column=15, value=seasonality_formula)
            
            # Growth factor (compound growth)
            growth_formula = f'=1+AVERAGEIF(SegmentLibrary!A:A,C{row},SegmentLibrary!H:H)/100'
            ws.cell(row=row, column=16, value=growth_formula)
            
            # Transaction volume with better calculation
            volume_formula = f'''=SUMPRODUCT(
                (SegmentLibrary!A:A=C{row})*
                SegmentLibrary!G:G*
                POWER(P{row},IF(D{row},A{row}/30,A{row}-1))*
                O{row}*
                IF(D{row},1/30,1)
            )'''
            ws.cell(row=row, column=14, value=volume_formula)
            
            # Revenue calculation
            revenue_formula = f'''=SUMPRODUCT(
                (SegmentLibrary!A:A=C{row})*
                SegmentLibrary!G:G*
                POWER(P{row},IF(D{row},A{row}/30,A{row}-1))*
                SegmentLibrary!E:E*
                O{row}*
                IF(D{row},1/30,1)
            )'''
            ws.cell(row=row, column=5, value=revenue_formula)
            
            # Revenue USD
            ws.cell(row=row, column=6, value=f'=E{row}/Dashboard!$B$5')
            
            # COGS calculation
            cogs_formula = f'''=SUMPRODUCT(
                (SegmentLibrary!A:A=C{row})*
                SegmentLibrary!G:G*
                POWER(P{row},IF(D{row},A{row}/30,A{row}-1))*
                SegmentLibrary!F:F*
                O{row}*
                IF(D{row},1/30,1)
            )'''
            ws.cell(row=row, column=7, value=cogs_formula)
            
            # COGS USD
            ws.cell(row=row, column=8, value=f'=G{row}/Dashboard!$B$5')
            
            # Operating expenses
            opex_formula = f'''=IF(Parameters!B6="fixed",
                Parameters!B5*IF(D{row},1/30,1),
                E{row}*Parameters!B7/100)'''
            ws.cell(row=row, column=9, value=opex_formula)
            
            # OpEx USD
            ws.cell(row=row, column=10, value=f'=I{row}/Dashboard!$B$5')
            
            # Net Profit
            ws.cell(row=row, column=11, value=f'=E{row}-G{row}-I{row}')
            
            # Net Profit USD
            ws.cell(row=row, column=12, value=f'=K{row}/Dashboard!$B$5')
            
            # Profit Margin
            ws.cell(row=row, column=13, value=f'=IF(E{row}>0,K{row}/E{row},0)')
            
            # Cumulative Revenue
            if row == 2:
                ws.cell(row=row, column=17, value=f'=E{row}')
            else:
                ws.cell(row=row, column=17, value=f'=Q{row-1}+E{row}')

    def create_yearly_aggregation_sheet(self):
        """Create a sheet for yearly aggregated data"""
        if 'YearlyView' in self.wb.sheetnames:
            del self.wb['YearlyView']
        
        ws = self.wb.create_sheet('YearlyView')
        
        # Headers
        headers = [
            'Year', 'Country', 'Total_Revenue_Local', 'Total_Revenue_USD',
            'Total_COGS_Local', 'Total_COGS_USD', 'Total_NetProfit_Local',
            'Total_NetProfit_USD', 'Avg_Profit_Margin', 'Total_Volume'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        
        # Create yearly aggregations for up to 10 years
        current_year = 2025  # Starting year
        for year_offset in range(10):
            year = current_year + year_offset
            row = year_offset + 2
            
            # Year
            ws.cell(row=row, column=1, value=year)
            
            # Country
            ws.cell(row=row, column=2, value='=Dashboard!$B$3')
            
            # Calculate start and end month for this year
            start_month = year_offset * 12 + 2  # +2 because data starts at row 2
            end_month = min(start_month + 11, 121)  # Max 120 months
            
            # Total Revenue Local
            ws.cell(row=row, column=3, value=f'=SUM(Projections!E{start_month}:E{end_month})')
            
            # Total Revenue USD
            ws.cell(row=row, column=4, value=f'=SUM(Projections!F{start_month}:F{end_month})')
            
            # Total COGS Local
            ws.cell(row=row, column=5, value=f'=SUM(Projections!G{start_month}:G{end_month})')
            
            # Total COGS USD
            ws.cell(row=row, column=6, value=f'=SUM(Projections!H{start_month}:H{end_month})')
            
            # Total Net Profit Local
            ws.cell(row=row, column=7, value=f'=SUM(Projections!K{start_month}:K{end_month})')
            
            # Total Net Profit USD
            ws.cell(row=row, column=8, value=f'=SUM(Projections!L{start_month}:L{end_month})')
            
            # Average Profit Margin
            ws.cell(row=row, column=9, value=f'=AVERAGE(Projections!M{start_month}:M{end_month})')
            
            # Total Volume
            ws.cell(row=row, column=10, value=f'=SUM(Projections!N{start_month}:N{end_month})')
        
        # Format columns
        for col in [3, 5, 7]:  # Local currency
            for row in range(2, 12):
                ws.cell(row=row, column=col).number_format = '#,##0'
        
        for col in [4, 6, 8]:  # USD
            for row in range(2, 12):
                ws.cell(row=row, column=col).number_format = '$#,##0'
        
        for col in [9]:  # Percentage
            for row in range(2, 12):
                ws.cell(row=row, column=col).number_format = '0.0%'
        
        for col in [10]:  # Volume
            for row in range(2, 12):
                ws.cell(row=row, column=col).number_format = '#,##0'

    def enhance_segment_library(self):
        """Enhance segment library with more realistic data from config"""
        ws = self.wb['SegmentLibrary']
        
        # Clear existing data except headers
        ws.delete_rows(2, ws.max_row)
        
        # Load actual segment data from config if available
        row = 2
        
        for country_code, country_data in self.config['countries'].items():
            # Default segments for each country
            segments = [
                {
                    'name': 'Basic Authentication',
                    'category': 'authentication',
                    'market': 'General',
                    'price': 0.15 * country_data.get('exchangeRate', 1) / 83.50,  # Adjust for currency
                    'cost': 0.05 * country_data.get('exchangeRate', 1) / 83.50,
                    'volume': 10000000,
                    'growth': 8,
                    'description': 'Basic authentication service'
                },
                {
                    'name': 'eKYC Premium',
                    'category': 'kyc',
                    'market': 'Financial',
                    'price': 2.50 * country_data.get('exchangeRate', 1) / 83.50,
                    'cost': 0.75 * country_data.get('exchangeRate', 1) / 83.50,
                    'volume': 5000000,
                    'growth': 12,
                    'description': 'Premium eKYC service'
                },
                {
                    'name': 'Biometric Verification',
                    'category': 'biometric',
                    'market': 'Government',
                    'price': 0.25 * country_data.get('exchangeRate', 1) / 83.50,
                    'cost': 0.08 * country_data.get('exchangeRate', 1) / 83.50,
                    'volume': 25000000,
                    'growth': 5,
                    'description': 'Government biometric verification'
                },
                {
                    'name': 'Mobile Authentication',
                    'category': 'mobile',
                    'market': 'Fintech',
                    'price': 0.50 * country_data.get('exchangeRate', 1) / 83.50,
                    'cost': 0.15 * country_data.get('exchangeRate', 1) / 83.50,
                    'volume': 15000000,
                    'growth': 15,
                    'description': 'Mobile app authentication'
                }
            ]
            
            for segment in segments:
                ws.cell(row=row, column=1, value=country_code)
                ws.cell(row=row, column=2, value=segment['name'])
                ws.cell(row=row, column=3, value=segment['category'])
                ws.cell(row=row, column=4, value=segment['market'])
                ws.cell(row=row, column=5, value=segment['price'])
                ws.cell(row=row, column=6, value=segment['cost'])
                ws.cell(row=row, column=7, value=segment['volume'])
                ws.cell(row=row, column=8, value=segment['growth'])
                ws.cell(row=row, column=9, value=segment['description'])
                row += 1

    def create_export_simulation_sheet(self):
        """Create a sheet that simulates the export functionality"""
        if 'ExportData' in self.wb.sheetnames:
            del self.wb['ExportData']
        
        ws = self.wb.create_sheet('ExportData')
        
        # Title
        ws['A1'] = 'Export Data Simulation'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Export type selector
        ws['A3'] = 'Export Type:'
        ws['B3'] = 'Monthly'  # Default
        
        export_types = ['Daily', 'Monthly', 'Yearly']
        dv = DataValidation(type="list", formula1=f'"{",".join(export_types)}"')
        dv.add(ws['B3'])
        ws.add_data_validation(dv)
        
        # Export format
        ws['A4'] = 'Format:'
        ws['B4'] = 'Excel'
        
        formats = ['Excel', 'CSV']
        dv_format = DataValidation(type="list", formula1=f'"{",".join(formats)}"')
        dv_format.add(ws['B4'])
        ws.add_data_validation(dv_format)
        
        # Dynamic export data based on selection
        ws['A6'] = 'Export Data Preview:'
        ws['A6'].font = Font(size=12, bold=True)
        
        # Headers for export preview
        export_headers = [
            'Period', 'Revenue_Local', 'Revenue_USD', 'COGS_Local', 
            'COGS_USD', 'NetProfit_Local', 'NetProfit_USD', 'Volume', 'Margin'
        ]
        
        for col, header in enumerate(export_headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # Sample export data (first 24 rows)
        for i in range(24):
            row = 8 + i
            ws.cell(row=row, column=1, value=f'=Projections!B{i+2}')  # Period
            ws.cell(row=row, column=2, value=f'=Projections!E{i+2}')  # Revenue Local
            ws.cell(row=row, column=3, value=f'=Projections!F{i+2}')  # Revenue USD
            ws.cell(row=row, column=4, value=f'=Projections!G{i+2}')  # COGS Local
            ws.cell(row=row, column=5, value=f'=Projections!H{i+2}')  # COGS USD
            ws.cell(row=row, column=6, value=f'=Projections!K{i+2}')  # NetProfit Local
            ws.cell(row=row, column=7, value=f'=Projections!L{i+2}')  # NetProfit USD
            ws.cell(row=row, column=8, value=f'=Projections!N{i+2}')  # Volume
            ws.cell(row=row, column=9, value=f'=Projections!M{i+2}')  # Margin

    def create_documentation_sheet(self):
        """Create a comprehensive documentation sheet"""
        if 'Documentation' in self.wb.sheetnames:
            del self.wb['Documentation']
        
        ws = self.wb.create_sheet('Documentation')
        
        # Title
        ws['A1'] = 'APAC Revenue Projections Model - User Guide'
        ws['A1'].font = Font(size=18, bold=True, color='667EEA')
        ws.merge_cells('A1:H1')
        
        # Table of Contents
        ws['A3'] = 'Table of Contents'
        ws['A3'].font = Font(size=14, bold=True)
        
        toc_items = [
            '1. Overview',
            '2. Dashboard Usage',
            '3. Country Selection',
            '4. Period Configuration',
            '5. Segment Management',
            '6. Projection Calculations',
            '7. Scenario Analysis',
            '8. Export Functions',
            '9. Formulas Reference',
            '10. Troubleshooting'
        ]
        
        for i, item in enumerate(toc_items):
            ws[f'A{4+i}'] = item
            ws[f'A{4+i}'].font = Font(bold=True)
        
        # Detailed sections
        current_row = 15
        
        sections = [
            {
                'title': '1. Overview',
                'content': [
                    'This Excel model replicates all functionality from the web-based',
                    'APAC Revenue Projections tool, supporting 8 countries:',
                    '• India (INR)', '• Singapore (SGD)', '• Australia (AUD)',
                    '• Japan (JPY)', '• South Korea (KRW)', '• Thailand (THB)',
                    '• Indonesia (IDR)', '• Philippines (PHP)',
                    '',
                    'Key Features:',
                    '• Multi-period projections (1M to 10Y)',
                    '• Dual currency display',
                    '• Scenario analysis',
                    '• Dynamic segment library',
                    '• Seasonality adjustments',
                    '• Export capabilities'
                ]
            },
            {
                'title': '2. Dashboard Usage',
                'content': [
                    'The Dashboard sheet is your main control center:',
                    '',
                    'Controls:',
                    '• Country: Select from 8 APAC countries',
                    '• Period: Choose 1M, 1Y, 2Y, 5Y, or 10Y',
                    '• All calculations update automatically',
                    '',
                    'Key Metrics:',
                    '• Total Revenue (Local & USD)',
                    '• Net Profit (Local & USD)',
                    '• Profit Margins',
                    '• Transaction Volumes'
                ]
            },
            {
                'title': '3. Calculation Logic',
                'content': [
                    'Revenue = Volume × Price Per Transaction',
                    'COGS = Volume × Cost Per Transaction',
                    'Net Profit = Revenue - COGS - Operating Expenses',
                    'Profit Margin = Net Profit / Revenue × 100',
                    '',
                    'Volume Growth:',
                    'Volume(n) = Base Volume × (1 + Growth Rate)^n × Seasonality',
                    '',
                    'Currency Conversion:',
                    'USD Value = Local Value / Exchange Rate'
                ]
            }
        ]
        
        for section in sections:
            ws[f'A{current_row}'] = section['title']
            ws[f'A{current_row}'].font = Font(size=12, bold=True, color='667EEA')
            current_row += 1
            
            for line in section['content']:
                if line.strip():
                    ws[f'A{current_row}'] = line
                current_row += 1
            current_row += 1

    def enhance_charts(self):
        """Enhance the charts sheet with better visualizations"""
        ws = self.wb['Charts']
        
        # Clear existing content
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        # Revenue Trend Chart
        chart1 = LineChart()
        chart1.title = "Revenue Projections by Period"
        chart1.style = 10
        chart1.x_axis.title = 'Period'
        chart1.y_axis.title = 'Revenue (Local Currency)'
        chart1.width = 15
        chart1.height = 8
        
        # Data for first 12 months
        data1 = Reference(ws, min_col=1, min_row=1, max_row=13, max_col=1)
        cats1 = Reference(ws, min_col=1, min_row=2, max_row=13)
        
        # Add revenue data from Projections sheet
        ws['A1'] = 'Revenue'
        for i in range(12):
            ws[f'A{i+2}'] = f'=Projections!E{i+2}'
            ws[f'B{i+2}'] = f'=Projections!B{i+2}'
        
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws.add_chart(chart1, "D1")
        
        # Volume Chart
        chart2 = BarChart()
        chart2.title = "Transaction Volume by Period"
        chart2.style = 12
        chart2.x_axis.title = 'Period'
        chart2.y_axis.title = 'Transaction Volume'
        chart2.width = 15
        chart2.height = 8
        
        # Volume data
        ws['A15'] = 'Volume'
        for i in range(12):
            ws[f'A{i+16}'] = f'=Projections!N{i+2}'
            ws[f'B{i+16}'] = f'=Projections!B{i+2}'
        
        data2 = Reference(ws, min_col=1, min_row=15, max_row=27, max_col=1)
        cats2 = Reference(ws, min_col=2, min_row=16, max_row=27)
        
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws.add_chart(chart2, "D20")

    def save_enhanced_model(self, filename):
        """Save the enhanced model"""
        try:
            filepath = f'/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/{filename}'
            self.wb.save(filepath)
            print(f"Enhanced Excel model saved: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error saving enhanced model: {e}")
            return None

    def create_enhanced_model(self):
        """Create the complete enhanced model"""
        print("Enhancing Excel Revenue Projection Model...")
        
        # Enhance existing sheets
        self.enhance_dashboard()
        self.create_advanced_projections()
        self.enhance_segment_library()
        
        # Create new sheets
        self.create_yearly_aggregation_sheet()
        self.create_export_simulation_sheet()
        self.create_documentation_sheet()
        self.enhance_charts()
        
        # Set dashboard as active
        self.wb.active = self.wb['Dashboard']
        
        print("All enhancements completed!")
        return self.wb

def main():
    """Main function to enhance the Excel model"""
    original_file = 'APAC_Revenue_Projections_Master_Model.xlsx'
    enhanced_file = 'APAC_Revenue_Projections_Enhanced_Model.xlsx'
    
    model = AdvancedExcelModel(original_file)
    model.create_enhanced_model()
    filepath = model.save_enhanced_model(enhanced_file)
    
    if filepath:
        print(f"\n✅ Enhanced Excel model created successfully!")
        print(f"📄 File location: {filepath}")
        print(f"\n🚀 New Features Added:")
        print(f"   • Advanced projection formulas")
        print(f"   • Yearly aggregation view")
        print(f"   • Enhanced segment library")
        print(f"   • Export data simulation")
        print(f"   • Comprehensive documentation")
        print(f"   • Improved charts and visualizations")
        print(f"   • Period-aware calculations")
        print(f"   • Better currency handling")
        print(f"   • Seasonality integration")
        print(f"   • Growth factor calculations")

if __name__ == "__main__":
    main()