#!/usr/bin/env python3
"""
Enhance Excel Model with Comprehensive Demographic Data
Adds demographic calculation sheets and integrates enhanced demographic data for all APAC countries
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import json
import os

def load_demographic_data():
    """Load the enhanced demographic data from external files"""
    config_path = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/model-config.json'
    demographics_dir = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/demographics'
    
    with open(config_path, 'r') as f:
        config = json.load(f)
    
    countries = config.get('countries', {})
    demographic_config = config.get('demographicDataConfig', {})
    
    # Load external demographic data if configured
    regional_data = {}
    if demographic_config.get('externalFiles') and os.path.exists(demographics_dir):
        print("Loading demographic data from external files...")
        
        # Load index file to get list of available countries
        index_path = os.path.join(demographics_dir, 'index.json')
        if os.path.exists(index_path):
            with open(index_path, 'r') as f:
                index_data = json.load(f)
            
            # Load each country's demographic data
            for country_info in index_data.get('countries', []):
                country_key = country_info['countryKey']
                file_name = country_info['fileName']
                file_path = os.path.join(demographics_dir, file_name)
                
                if os.path.exists(file_path):
                    with open(file_path, 'r') as f:
                        country_data = json.load(f)
                    
                    regional_data[country_key] = {
                        'demographicSegments': country_data.get('demographicSegments', [])
                    }
                    print(f"  ✅ Loaded {len(country_data.get('demographicSegments', []))} segments for {country_data['country']['name']}")
                else:
                    print(f"  ⚠️  File not found: {file_path}")
        else:
            print(f"  ⚠️  Index file not found: {index_path}")
    else:
        print("Using embedded demographic data from model-config.json...")
        regional_data = config.get('regionalData', {})
    
    return regional_data, countries

def create_demographic_summary_sheet(wb, regional_data, countries):
    """Create a summary sheet showing demographic overview for all countries"""
    
    if 'DemographicSummary' in wb.sheetnames:
        ws = wb['DemographicSummary']
        wb.remove(ws)
    
    ws = wb.create_sheet('DemographicSummary', 1)
    
    # Headers
    headers = ['Country', 'Total Population (M)', 'Segments', 'Avg Auth Rate (%)', 
               'Avg Digital Adoption (%)', 'High Economic Tier (%)', 'Urban Population (%)']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Calculate summary data for each country
    row = 2
    for country_key, demographic_segments in regional_data.items():
        if 'demographicSegments' not in demographic_segments:
            continue
            
        segments = demographic_segments['demographicSegments']
        country_name = countries.get(country_key, {}).get('name', country_key.title())
        
        # Calculate aggregated statistics
        total_population = sum(seg.get('population', 0) for seg in segments)
        avg_auth_rate = sum(seg.get('authPct', 0) for seg in segments) / len(segments) if segments else 0
        avg_digital_adoption = sum(seg.get('digitalAdoption', 0) for seg in segments) / len(segments) if segments else 0
        high_economic_segments = [seg for seg in segments if seg.get('economicTier') == 'high']
        high_economic_pct = (len(high_economic_segments) / len(segments)) * 100 if segments else 0
        avg_urbanization = sum(seg.get('urbanization', 0) for seg in segments) / len(segments) if segments else 0
        
        # Add row data
        ws.cell(row=row, column=1, value=country_name)
        ws.cell(row=row, column=2, value=round(total_population, 1))
        ws.cell(row=row, column=3, value=len(segments))
        ws.cell(row=row, column=4, value=round(avg_auth_rate, 1))
        ws.cell(row=row, column=5, value=round(avg_digital_adoption, 1))
        ws.cell(row=row, column=6, value=round(high_economic_pct, 1))
        ws.cell(row=row, column=7, value=round(avg_urbanization, 1))
        
        row += 1
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    print("✅ Created Demographic Summary sheet")

def create_country_demographic_sheets(wb, regional_data, countries):
    """Create detailed demographic sheets for each country"""
    
    for country_key, demographic_data in regional_data.items():
        if 'demographicSegments' not in demographic_data:
            continue
            
        segments = demographic_data['demographicSegments']
        country_name = countries.get(country_key, {}).get('name', country_key.title())
        sheet_name = f'{country_name}_Demographics'
        
        # Remove existing sheet if it exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        
        ws = wb.create_sheet(sheet_name)
        
        # Headers for detailed demographic data
        headers = ['Segment Name', 'Population (M)', 'Auth Rate (%)', 'Auth Frequency', 
                   'Digital Adoption (%)', 'Economic Tier', 'Urbanization (%)', 
                   'Growth Rate (%)', 'Monthly Volume', 'Revenue Potential']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add demographic segment data
        row = 2
        for segment in segments:
            ws.cell(row=row, column=1, value=segment.get('name', ''))
            ws.cell(row=row, column=2, value=segment.get('population', 0))
            ws.cell(row=row, column=3, value=segment.get('authPct', 0))
            ws.cell(row=row, column=4, value=segment.get('authFreq', 1.0))
            ws.cell(row=row, column=5, value=segment.get('digitalAdoption', 0))
            ws.cell(row=row, column=6, value=segment.get('economicTier', 'medium'))
            ws.cell(row=row, column=7, value=segment.get('urbanization', 0))
            ws.cell(row=row, column=8, value=segment.get('authGrowthRate', 3))
            
            # Calculate monthly volume and revenue potential
            population_millions = segment.get('population', 0)
            auth_rate = segment.get('authPct', 0) / 100
            auth_frequency = segment.get('authFreq', 1.0)
            monthly_volume = population_millions * 1000000 * auth_rate * auth_frequency
            
            # Estimate revenue potential based on economic tier
            price_per_transaction = 0.12  # Default
            if segment.get('economicTier') == 'high':
                price_per_transaction = 0.18
            elif segment.get('economicTier') == 'low':
                price_per_transaction = 0.08
                
            # Digital adoption bonus
            if segment.get('digitalAdoption', 0) >= 80:
                price_per_transaction *= 1.2
            elif segment.get('digitalAdoption', 0) <= 50:
                price_per_transaction *= 0.9
            
            monthly_revenue_potential = monthly_volume * price_per_transaction
            
            ws.cell(row=row, column=9, value=int(monthly_volume))
            ws.cell(row=row, column=10, value=int(monthly_revenue_potential))
            
            row += 1
        
        # Add summary row
        if segments:
            summary_row = row + 1
            ws.cell(row=summary_row, column=1, value='TOTAL')
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            # Sum formulas
            ws.cell(row=summary_row, column=2, value=f'=SUM(B2:B{row-1})')  # Total population
            ws.cell(row=summary_row, column=9, value=f'=SUM(I2:I{row-1})')  # Total volume
            ws.cell(row=summary_row, column=10, value=f'=SUM(J2:J{row-1})')  # Total revenue potential
            
            # Average formulas
            ws.cell(row=summary_row, column=3, value=f'=AVERAGE(C2:C{row-1})')  # Avg auth rate
            ws.cell(row=summary_row, column=5, value=f'=AVERAGE(E2:E{row-1})')  # Avg digital adoption
            ws.cell(row=summary_row, column=7, value=f'=AVERAGE(G2:G{row-1})')  # Avg urbanization
            ws.cell(row=summary_row, column=8, value=f'=AVERAGE(H2:H{row-1})')  # Avg growth rate
            
            for col in range(1, 11):
                ws.cell(row=summary_row, column=col).font = Font(bold=True)
                ws.cell(row=summary_row, column=col).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Created detailed demographic sheet for {country_name}")

def enhance_dashboard_with_demographics(wb):
    """Enhance the Dashboard sheet with demographic selection options"""
    
    if 'Dashboard' not in wb.sheetnames:
        return
        
    ws = wb['Dashboard']
    
    # Add demographic analysis section
    demo_start_row = 12
    
    # Headers
    ws.cell(row=demo_start_row, column=1, value='Demographic Analysis')
    ws.cell(row=demo_start_row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=demo_start_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.cell(row=demo_start_row, column=1).font = Font(color='FFFFFF', bold=True, size=14)
    
    # Merge cells for header
    ws.merge_cells(f'A{demo_start_row}:C{demo_start_row}')
    
    # Add demographic parameters
    demo_params = [
        ('Selected Segment', 'All Segments'),
        ('Economic Tier Filter', 'All Tiers'),
        ('Min Digital Adoption (%)', 50),
        ('Min Urbanization (%)', 0),
        ('Revenue Multiplier', 1.0)
    ]
    
    row = demo_start_row + 1
    for param, default_value in demo_params:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=default_value)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
    
    # Add demographic summary formulas
    summary_start_row = row + 1
    ws.cell(row=summary_start_row, column=1, value='Demographic Insights')
    ws.cell(row=summary_start_row, column=1).font = Font(bold=True, size=12)
    
    insights = [
        ('Total Segments Available', f'=COUNTA(DemographicSummary.A2:A20)'),
        ('Avg Auth Rate (%)', f'=AVERAGE(DemographicSummary.D2:D20)'),
        ('Avg Digital Adoption (%)', f'=AVERAGE(DemographicSummary.E2:E20)'),
        ('High Economic Tier Countries', f'=COUNTIF(DemographicSummary.F2:F20,">50")')
    ]
    
    row = summary_start_row + 1
    for insight, formula in insights:
        ws.cell(row=row, column=1, value=insight)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=1).font = Font(italic=True)
        row += 1
    
    print("✅ Enhanced Dashboard with demographic analysis section")

def enhance_excel_model_with_demographics():
    """Main function to enhance the Excel model with comprehensive demographic data"""
    
    file_path = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/APAC_Revenue_Projections_Enhanced_Model.xlsx'
    
    if not os.path.exists(file_path):
        print("❌ Excel file not found")
        return False
    
    try:
        # Load demographic data
        regional_data, countries = load_demographic_data()
        
        # Open workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Create demographic sheets
        create_demographic_summary_sheet(wb, regional_data, countries)
        create_country_demographic_sheets(wb, regional_data, countries)
        enhance_dashboard_with_demographics(wb)
        
        # Save enhanced workbook
        wb.save(file_path)
        
        print(f"\n🎉 Excel model successfully enhanced with comprehensive demographic data!")
        print(f"📄 Enhanced file: {file_path}")
        print(f"\n📊 New Features Added:")
        print(f"   • Demographic Summary sheet with country comparisons")
        print(f"   • Individual demographic sheets for each APAC country")
        print(f"   • Enhanced demographic fields: digital adoption, economic tiers, urbanization")
        print(f"   • Growth rate calculations based on demographic data")
        print(f"   • Revenue potential calculations with demographic adjustments")
        print(f"   • Dashboard integration with demographic analysis")
        
        return True
        
    except Exception as e:
        print(f"❌ Error enhancing Excel model with demographics: {str(e)}")
        return False

if __name__ == "__main__":
    enhance_excel_model_with_demographics()