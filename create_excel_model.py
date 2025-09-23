#!/usr/bin/env python3
"""
Create a comprehensive Excel revenue projection model that contains all the logic
from the web-based tool, supporting all APAC countries and calculation methods.
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import math
from datetime import datetime, timedelta
import os

class ExcelRevenueModel:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Load configuration data
        self.load_config_data()
        
        # Style definitions
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        self.currency_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def load_config_data(self):
        """Load configuration data from model-config.json"""
        config_path = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/model-config.json'
        try:
            with open(config_path, 'r') as f:
                self.config = json.load(f)
        except FileNotFoundError:
            # Fallback configuration if file not found
            self.config = self.create_fallback_config()
    
    def create_fallback_config(self):
        """Create fallback configuration data"""
        return {
            "countries": {
                "india": {"name": "India", "currency": "INR", "currencySymbol": "₹", "exchangeRate": 83.50},
                "singapore": {"name": "Singapore", "currency": "SGD", "currencySymbol": "S$", "exchangeRate": 1.35},
                "australia": {"name": "Australia", "currency": "AUD", "currencySymbol": "A$", "exchangeRate": 1.52},
                "japan": {"name": "Japan", "currency": "JPY", "currencySymbol": "¥", "exchangeRate": 149.0},
                "south_korea": {"name": "South Korea", "currency": "KRW", "currencySymbol": "₩", "exchangeRate": 1320.0},
                "thailand": {"name": "Thailand", "currency": "THB", "currencySymbol": "฿", "exchangeRate": 35.8},
                "indonesia": {"name": "Indonesia", "currency": "IDR", "currencySymbol": "Rp", "exchangeRate": 15750.0},
                "philippines": {"name": "Philippines", "currency": "PHP", "currencySymbol": "₱", "exchangeRate": 56.2}
            },
            "segmentLibraries": {},
            "regionalData": {}
        }

    def create_dashboard_sheet(self):
        """Create the main dashboard sheet"""
        ws = self.wb.create_sheet(title="Dashboard")
        
        # Title and header
        ws['A1'] = 'APAC Revenue Projections Dashboard'
        ws['A1'].font = Font(size=18, bold=True, color='667EEA')
        ws.merge_cells('A1:H1')
        
        # Country selector
        ws['A3'] = 'Country:'
        ws['B3'] = 'india'  # Default
        ws['A3'].font = Font(bold=True)
        
        # Create country dropdown validation
        countries = list(self.config['countries'].keys())
        dv = DataValidation(type="list", formula1=f'"{",".join(countries)}"')
        dv.add(ws['B3'])
        ws.add_data_validation(dv)
        
        # Period selector
        ws['A4'] = 'Period:'
        ws['B4'] = '1Y'  # Default
        ws['A4'].font = Font(bold=True)
        
        periods = ['1M', '1Y', '2Y', '5Y', '10Y']
        dv_period = DataValidation(type="list", formula1=f'"{",".join(periods)}"')
        dv_period.add(ws['B4'])
        ws.add_data_validation(dv_period)
        
        # Exchange rate display
        ws['A5'] = 'Exchange Rate:'
        ws['B5'] = '=VLOOKUP(B3,CountryData!A:D,4,FALSE)'
        ws['A5'].font = Font(bold=True)
        
        # Currency symbol
        ws['A6'] = 'Currency:'
        ws['B6'] = '=VLOOKUP(B3,CountryData!A:C,3,FALSE)'
        ws['A6'].font = Font(bold=True)
        
        # Key metrics headers
        metrics_row = 8
        ws[f'A{metrics_row}'] = 'Key Metrics'
        ws[f'A{metrics_row}'].font = Font(size=14, bold=True)
        
        # Metric cards
        metrics = [
            ('Total Revenue (Local)', f'=SUM(Projections!D:D)'),
            ('Total Revenue (USD)', f'=SUM(Projections!D:D)/B5'),
            ('Total Net Profit (Local)', f'=SUM(Projections!H:H)'),
            ('Total Net Profit (USD)', f'=SUM(Projections!H:H)/B5'),
            ('Avg Profit Margin', f'=AVERAGE(Projections!I:I)'),
            ('Total Transaction Volume', f'=SUM(Projections!J:J)')
        ]
        
        for i, (label, formula) in enumerate(metrics):
            row = metrics_row + 2 + i
            ws[f'A{row}'] = label
            ws[f'B{row}'] = formula
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].number_format = '#,##0'
        
        # Auto-size columns
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 20
            
        return ws

    def create_country_data_sheet(self):
        """Create the country configuration data sheet"""
        ws = self.wb.create_sheet(title="CountryData")
        
        # Headers
        headers = ['CountryCode', 'CountryName', 'CurrencySymbol', 'ExchangeRate', 'Population']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Country data
        row = 2
        for code, data in self.config['countries'].items():
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=data.get('name', code.title()))
            ws.cell(row=row, column=3, value=data.get('currencySymbol', '$'))
            ws.cell(row=row, column=4, value=data.get('exchangeRate', 1))
            ws.cell(row=row, column=5, value=data.get('population', 0))
            row += 1
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        return ws

    def create_parameters_sheet(self):
        """Create the input parameters sheet"""
        ws = self.wb.create_sheet(title="Parameters")
        
        # Title
        ws['A1'] = 'Model Parameters'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Base parameters section
        ws['A3'] = 'Base Parameters'
        ws['A3'].font = Font(size=12, bold=True)
        
        params = [
            ('Start Revenue', 0),
            ('Growth Rate (%)', 4),
            ('Projection Months', 12),
            ('Cost Percentage (%)', 35),
            ('Operating Expenses', 0),
            ('Operating Expense Type', 'fixed'),
            ('Operating Expense Percentage (%)', 15)
        ]
        
        row = 4
        for param, default_value in params:
            ws[f'A{row}'] = param
            ws[f'B{row}'] = default_value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Seasonality section
        ws['A12'] = 'Seasonality Multipliers'
        ws['A12'].font = Font(size=12, bold=True)
        
        # Seasonality data
        seasonality_data = {
            'Month': list(range(1, 13)),
            'None': [1.0] * 12,
            'Retail': [0.9, 0.85, 0.9, 0.95, 1.0, 1.05, 1.1, 1.05, 1.0, 1.1, 1.2, 1.3],
            'Summer': [0.8, 0.85, 0.9, 1.0, 1.1, 1.2, 1.3, 1.2, 1.0, 0.9, 0.85, 0.8]
        }
        
        # Headers for seasonality table
        headers = ['Month', 'None', 'Retail', 'Summer']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Seasonality values
        for month in range(12):
            ws.cell(row=14 + month, column=1, value=month + 1)
            ws.cell(row=14 + month, column=2, value=seasonality_data['None'][month])
            ws.cell(row=14 + month, column=3, value=seasonality_data['Retail'][month])
            ws.cell(row=14 + month, column=4, value=seasonality_data['Summer'][month])
        
        return ws

    def create_segments_sheet(self):
        """Create the segment library sheet"""
        ws = self.wb.create_sheet(title="SegmentLibrary")
        
        # Headers
        headers = [
            'Country', 'SegmentName', 'Category', 'Market', 
            'PricePerTransaction', 'CostPerTransaction', 
            'MonthlyVolume', 'VolumeGrowth', 'Description'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Sample segment data for each country
        segments_data = []
        
        for country_code in self.config['countries'].keys():
            country_segments = [
                {
                    'country': country_code,
                    'name': 'Basic Authentication',
                    'category': 'authentication',
                    'market': 'General',
                    'price': 0.15,
                    'cost': 0.05,
                    'volume': 10000000,
                    'growth': 8,
                    'description': 'Basic authentication service for general verification'
                },
                {
                    'country': country_code,
                    'name': 'eKYC Service',
                    'category': 'kyc',
                    'market': 'Financial Services',
                    'price': 2.50,
                    'cost': 0.75,
                    'volume': 5000000,
                    'growth': 12,
                    'description': 'Electronic KYC service for financial institutions'
                },
                {
                    'country': country_code,
                    'name': 'Biometric Auth',
                    'category': 'biometric',
                    'market': 'Government',
                    'price': 0.25,
                    'cost': 0.08,
                    'volume': 25000000,
                    'growth': 5,
                    'description': 'Biometric authentication for government services'
                }
            ]
            segments_data.extend(country_segments)
        
        # Add segment data to sheet
        row = 2
        for segment in segments_data:
            ws.cell(row=row, column=1, value=segment['country'])
            ws.cell(row=row, column=2, value=segment['name'])
            ws.cell(row=row, column=3, value=segment['category'])
            ws.cell(row=row, column=4, value=segment['market'])
            ws.cell(row=row, column=5, value=segment['price'])
            ws.cell(row=row, column=6, value=segment['cost'])
            ws.cell(row=row, column=7, value=segment['volume'])
            ws.cell(row=row, column=8, value=segment['growth'])
            ws.cell(row=row, column=9, value=segment['description'])
            row += 1
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column].width = adjusted_width
            
        return ws

    def create_projections_sheet(self):
        """Create the main projections calculation sheet"""
        ws = self.wb.create_sheet(title="Projections")
        
        # Headers
        headers = [
            'Month', 'Period', 'Country', 'Revenue_Local', 'Revenue_USD',
            'COGS_Local', 'COGS_USD', 'NetProfit_Local', 'NetProfit_USD',
            'ProfitMargin', 'TransactionVolume'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Create formulas for 120 months (10 years max)
        for month in range(1, 121):
            row = month + 1
            
            # Month number
            ws.cell(row=row, column=1, value=month)
            
            # Period calculation based on dashboard selection
            period_formula = f'=IF(Dashboard!B4="1M",TEXT(DATE(YEAR(TODAY()),MONTH(TODAY())+{month-1},1),"mmm dd"),' \
                           f'TEXT(DATE(YEAR(TODAY()),MONTH(TODAY())+{month-1},1),"yyyy mmm"))'
            ws.cell(row=row, column=2, value=period_formula)
            
            # Country reference
            ws.cell(row=row, column=3, value='=Dashboard!B3')
            
            # Revenue calculation (sum of all segments for this country/month)
            revenue_formula = f'=SUMPRODUCT(' \
                            f'(SegmentLibrary!A:A=C{row}),' \
                            f'SegmentLibrary!G:G*POWER(1+SegmentLibrary!H:H/100,A{row}-1),' \
                            f'SegmentLibrary!E:E,' \
                            f'INDEX(Parameters!B:D,MOD(A{row}-1,12)+1,2))'
            ws.cell(row=row, column=4, value=revenue_formula)
            
            # Revenue USD
            ws.cell(row=row, column=5, value=f'=D{row}/Dashboard!B5')
            
            # COGS calculation
            cogs_formula = f'=SUMPRODUCT(' \
                         f'(SegmentLibrary!A:A=C{row}),' \
                         f'SegmentLibrary!G:G*POWER(1+SegmentLibrary!H:H/100,A{row}-1),' \
                         f'SegmentLibrary!F:F,' \
                         f'INDEX(Parameters!B:D,MOD(A{row}-1,12)+1,2))'
            ws.cell(row=row, column=6, value=cogs_formula)
            
            # COGS USD
            ws.cell(row=row, column=7, value=f'=F{row}/Dashboard!B5')
            
            # Net Profit (Revenue - COGS - Operating Expenses)
            ws.cell(row=row, column=8, value=f'=D{row}-F{row}-Parameters!B6')
            
            # Net Profit USD
            ws.cell(row=row, column=9, value=f'=H{row}/Dashboard!B5')
            
            # Profit Margin
            ws.cell(row=row, column=10, value=f'=IF(D{row}>0,H{row}/D{row}*100,0)')
            
            # Transaction Volume
            volume_formula = f'=SUMPRODUCT(' \
                           f'(SegmentLibrary!A:A=C{row}),' \
                           f'SegmentLibrary!G:G*POWER(1+SegmentLibrary!H:H/100,A{row}-1),' \
                           f'INDEX(Parameters!B:D,MOD(A{row}-1,12)+1,2))'
            ws.cell(row=row, column=11, value=volume_formula)
        
        # Format currency columns
        for col in [4, 6, 8]:  # Local currency columns
            for row in range(2, 122):
                ws.cell(row=row, column=col).number_format = '#,##0'
        
        for col in [5, 7, 9]:  # USD columns
            for row in range(2, 122):
                ws.cell(row=row, column=col).number_format = '$#,##0'
        
        for col in [10]:  # Percentage columns
            for row in range(2, 122):
                ws.cell(row=row, column=col).number_format = '0.0%'
        
        for col in [11]:  # Volume columns
            for row in range(2, 122):
                ws.cell(row=row, column=col).number_format = '#,##0'
        
        return ws

    def create_scenarios_sheet(self):
        """Create scenario analysis sheet"""
        ws = self.wb.create_sheet(title="Scenarios")
        
        # Title
        ws['A1'] = 'Scenario Analysis'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Scenario parameters
        scenarios = [
            ('Conservative', 0.7, 0.9, 1.1, 1.0),
            ('Base Case', 1.0, 1.0, 1.0, 1.0),
            ('Optimistic', 1.5, 1.1, 0.9, 1.0)
        ]
        
        # Headers
        headers = ['Scenario', 'Volume Multiplier', 'Price Multiplier', 'Cost Multiplier', 'OpEx Multiplier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Scenario data
        for row, (name, vol_mult, price_mult, cost_mult, opex_mult) in enumerate(scenarios, 4):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=vol_mult)
            ws.cell(row=row, column=3, value=price_mult)
            ws.cell(row=row, column=4, value=cost_mult)
            ws.cell(row=row, column=5, value=opex_mult)
        
        # Results section
        ws['A8'] = 'Scenario Results'
        ws['A8'].font = Font(size=14, bold=True)
        
        result_headers = ['Scenario', 'Total Revenue', 'Total Profit', 'Profit Margin', 'Avg Monthly Revenue']
        for col, header in enumerate(result_headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Calculate results for each scenario
        for row, (name, vol_mult, price_mult, cost_mult, opex_mult) in enumerate(scenarios, 10):
            ws.cell(row=row, column=1, value=name)
            
            # Total Revenue with scenario adjustments
            revenue_formula = f'=SUMPRODUCT(Projections!D:D)*{vol_mult}*{price_mult}'
            ws.cell(row=row, column=2, value=revenue_formula)
            
            # Total Profit with scenario adjustments
            profit_formula = f'=(SUMPRODUCT(Projections!D:D)*{vol_mult}*{price_mult})' \
                           f'-(SUMPRODUCT(Projections!F:F)*{vol_mult}*{cost_mult})' \
                           f'-(Parameters!B6*{opex_mult}*Parameters!B3)'
            ws.cell(row=row, column=3, value=profit_formula)
            
            # Profit Margin
            ws.cell(row=row, column=4, value=f'=IF(B{row}>0,C{row}/B{row}*100,0)')
            
            # Average Monthly Revenue
            ws.cell(row=row, column=5, value=f'=B{row}/Parameters!B3')
        
        return ws

    def create_charts_sheet(self):
        """Create charts and visualizations sheet"""
        ws = self.wb.create_sheet(title="Charts")
        
        # Revenue projection chart
        chart = LineChart()
        chart.title = "Revenue Projections"
        chart.style = 10
        chart.x_axis.title = 'Month'
        chart.y_axis.title = 'Revenue'
        
        # Data for chart (first 12 months)
        data = Reference(ws, min_col=4, min_row=1, max_row=13, max_col=4)
        cats = Reference(ws, min_col=2, min_row=2, max_row=13)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "A1")
        
        return ws

    def apply_styling(self):
        """Apply consistent styling across all sheets"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            # Apply borders to data ranges
            if hasattr(ws, 'max_row') and ws.max_row > 1:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                      min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.border = self.border

    def save_workbook(self, filename):
        """Save the workbook to file"""
        try:
            filepath = f'/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/{filename}'
            self.wb.save(filepath)
            print(f"Excel model saved successfully: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return None

    def create_complete_model(self):
        """Create the complete Excel model with all sheets"""
        print("Creating Excel Revenue Projection Model...")
        
        # Create all sheets
        self.create_country_data_sheet()
        self.create_parameters_sheet()
        self.create_segments_sheet()
        self.create_projections_sheet()
        self.create_scenarios_sheet()
        self.create_dashboard_sheet()
        self.create_charts_sheet()
        
        # Apply styling
        self.apply_styling()
        
        # Set dashboard as active sheet
        self.wb.active = self.wb['Dashboard']
        
        print("All sheets created successfully!")
        
        return self.wb

def main():
    """Main function to create the Excel model"""
    model = ExcelRevenueModel()
    model.create_complete_model()
    filepath = model.save_workbook('APAC_Revenue_Projections_Master_Model.xlsx')
    
    if filepath:
        print(f"\n✅ Excel model created successfully!")
        print(f"📄 File location: {filepath}")
        print(f"\n📊 Model Features:")
        print(f"   • 8 APAC Countries supported")
        print(f"   • Multi-period projections (1M to 10Y)")
        print(f"   • Dual currency display (Local + USD)")
        print(f"   • Scenario analysis (Conservative/Base/Optimistic)")
        print(f"   • Segment library with authentication services")
        print(f"   • Dynamic dashboard with country/period selection")
        print(f"   • Seasonality adjustments")
        print(f"   • Volume growth calculations")
        print(f"   • Comprehensive financial modeling")

if __name__ == "__main__":
    main()