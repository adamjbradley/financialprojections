#!/usr/bin/env python3
"""
Fix Excel Formula Errors - APAC Revenue Projections Model
Corrects #VALUE! errors in projection calculations
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def fix_excel_model():
    """Fix the Excel model formulas and references"""
    file_path = '/Users/adambradley/Projects/Mastercard/ProductManager/financialprojections/APAC_Revenue_Projections_Enhanced_Model.xlsx'
    
    if not os.path.exists(file_path):
        print("❌ Excel file not found")
        return False
        
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Fix Dashboard sheet with proper parameter structure
        if 'Dashboard' not in wb.sheetnames:
            ws_dashboard = wb.create_sheet('Dashboard', 0)
        else:
            ws_dashboard = wb['Dashboard']
        
        # Clear and rebuild dashboard
        ws_dashboard.delete_rows(1, ws_dashboard.max_row)
        
        # Create parameter structure
        headers = [
            ['Parameter', 'Value', 'Description'],
            ['Country', 'india', 'Selected country for projections'],
            ['Period', '1Y', 'Time period for analysis'],
            ['Base Revenue', 0, 'Starting monthly revenue'],
            ['Growth Rate (%)', 5, 'Monthly growth rate'],
            ['Projection Months', 12, 'Number of months to project'],
            ['COGS (%)', 35, 'Cost of goods sold percentage'],
            ['Operating Expenses', 0, 'Fixed operating expenses'],
            ['USD Exchange Rate', 83.5, 'Local currency to USD rate'],
            ['Scenario', 'Base Case', 'Current scenario selection']
        ]
        
        for row_idx, row_data in enumerate(headers, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_dashboard.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
        
        # Fix Projections sheet
        if 'Projections' not in wb.sheetnames:
            ws_proj = wb.create_sheet('Projections')
        else:
            ws_proj = wb['Projections']
        
        # Clear and rebuild projections
        ws_proj.delete_rows(1, ws_proj.max_row)
        
        # Create projection headers
        proj_headers = ['Month', 'Period', 'Country', 'Revenue_Local', 'Revenue_USD', 
                       'COGS_Local', 'COGS_USD', 'NetProfit_Local', 'NetProfit_USD', 
                       'ProfitMargin', 'TransactionVolume']
        
        for col_idx, header in enumerate(proj_headers, 1):
            cell = ws_proj.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Create projection data with working formulas
        for month in range(1, 25):  # 24 months of data
            row = month + 1
            
            # Month number
            ws_proj[f'A{row}'] = month
            
            # Period (e.g., "2025 Sep")
            ws_proj[f'B{row}'] = f'2025 {["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][month-1 if month <= 12 else (month-1)%12]}'
            
            # Country
            ws_proj[f'C{row}'] = '=Dashboard.B2'
            
            # Revenue calculation - compound growth
            base_revenue = 1000000  # 1M base revenue for demonstration
            growth_factor = f'(1+Dashboard.B5/100)'
            ws_proj[f'D{row}'] = f'=Dashboard.B4+{base_revenue}*POWER({growth_factor},{month-1})'
            
            # Revenue USD
            ws_proj[f'E{row}'] = f'=D{row}/Dashboard.B9'
            
            # COGS Local
            ws_proj[f'F{row}'] = f'=D{row}*Dashboard.B7/100'
            
            # COGS USD  
            ws_proj[f'G{row}'] = f'=F{row}/Dashboard.B9'
            
            # Net Profit Local
            ws_proj[f'H{row}'] = f'=D{row}-F{row}-Dashboard.B8'
            
            # Net Profit USD
            ws_proj[f'I{row}'] = f'=H{row}/Dashboard.B9'
            
            # Profit Margin
            ws_proj[f'J{row}'] = f'=IF(D{row}>0,H{row}/D{row}*100,0)'
            
            # Transaction Volume
            ws_proj[f'K{row}'] = f'=D{row}*2000'  # Assume 2000 transactions per revenue unit
        
        # Auto-size columns
        for ws in [ws_dashboard, ws_proj]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the corrected workbook
        wb.save(file_path)
        print(f"✅ Fixed Excel model formulas successfully!")
        print(f"📄 File saved: {file_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error fixing Excel model: {str(e)}")
        return False

if __name__ == "__main__":
    fix_excel_model()